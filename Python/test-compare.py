import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

input_path = r'E:\WorkAndStudy\Python\result-cn\result_battlepass.xlsx'
output_path = r'E:\WorkAndStudy\Python\result-cn\result_battlepass-color.xlsx'
sheet = 'bplevel'
client1 = 'D'
server1 = 'I'
client2 = 'N'
server2 = 'S'

#取左边数
def ClientValLeft(strval):
    print('str:{}, clientvalleft:{}'.format(strval, strval.split('|')[0]))
    return strval.split('|')[0]
def ServerValLeft(strval):
    print('str:{}, servervalleft:{}'.format(strval, strval.split(':')[0]))
    return strval.split(':')[0]

#取右边数
def ClientValRight(strval):
    print('str:{}, clientvalright:{}'.format(strval, strval.split('|')[1]))
    return strval.split('|')[1]
def ServerValRight(strval):
    print('str:{}, servervalright:{}'.format(strval, strval.split(':')[1]))
    return strval.split(':')[1]

def ABCto123(strval):
    return ord(strval)-ord('A')

def color_spray(input_path, output_path, sheet, client1, server1, client2, server2):
    xlsx_data = pd.read_excel(input_path, sheet_name=sheet)
    col1_name = xlsx_data.columns[ABCto123(client1)]
    col2_name = xlsx_data.columns[ABCto123(server1)]
    col1_data = xlsx_data[col1_name]
    col2_data = xlsx_data[col2_name]
    col3_name = xlsx_data.columns[ABCto123(client2)]
    col4_name = xlsx_data.columns[ABCto123(server2)]
    col3_data = xlsx_data[col3_name]
    col4_data = xlsx_data[col4_name]

    #获取免费2列中左值不同的行
    diff_lines1 = []
    for i in range(len(col1_data.values)):
        if ClientValLeft(str(col1_data.values[i])) != ServerValLeft(str(col2_data.values[i])):
            diff_lines1.append(i)
            print(diff_lines1)
    #获取免费2列中右值不同的行
    diff_lines2 = []
    for i in range(len(col1_data.values)):
        if ClientValRight(str(col1_data.values[i])) != ServerValRight(str(col2_data.values[i])):
            diff_lines2.append(i)
            print(diff_lines2)
    #获取付费2列中左值不同的行
    diff_lines3 = []
    for i in range(len(col3_data.values)):
        if ClientValLeft(str(col3_data.values[i])) != ServerValLeft(str(col4_data.values[i])):
            diff_lines3.append(i)
            print(diff_lines3)
    # 获取付费2列中右值不同的行
    diff_lines4 = []
    for i in range(len(col3_data.values)):
        if ClientValRight(str(col3_data.values[i])) != ServerValRight(str(col4_data.values[i])):
            diff_lines4.append(i)
            print(diff_lines4)

    """上色"""
    wb = load_workbook(filename=input_path)
    # 使用第一个sheet作为工作簿
    work = wb[sheet]
    fill1 = PatternFill('solid', fgColor='fff68f')
    fill2 = PatternFill('solid', fgColor='a65fff')
    fill3 = PatternFill('solid', fgColor='ff5fb8')
    fill4 = PatternFill('solid', fgColor='ff715f')


    for i in diff_lines1:
        # work[i].fill = fill1
        work.cell(row=i + 2, column=ABCto123(client1) + 1).fill = fill1
        work.cell(row=i + 2, column=ABCto123(server1) + 1).fill = fill1
    for i in diff_lines2:
        # work[i].fill = fill1
        work.cell(row=i + 2, column=ABCto123(client1) + 1).fill = fill2
        work.cell(row=i + 2, column=ABCto123(server1) + 1).fill = fill2
    for i in diff_lines3:
        # work[i].fill = fill1
        work.cell(row=i + 2, column=ABCto123(client2) + 1).fill = fill3
        work.cell(row=i + 2, column=ABCto123(server2) + 1).fill = fill3
    for i in diff_lines4:
        # work[i].fill = fill1
        work.cell(row=i + 2, column=ABCto123(client2) + 1).fill = fill4
        work.cell(row=i + 2, column=ABCto123(server2) + 1).fill = fill4

    wb.close()
    wb.save(output_path)

color_spray(input_path, output_path, sheet, client1, server1, client2, server2)

