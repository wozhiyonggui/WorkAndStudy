import sys
import pandas as pd
import  argparse
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

#命令行解析器对象
parser = argparse.ArgumentParser(description='Demo of argparse')

#命令行参数
parser.add_argument('--path_config_battlepass', type=str, default=r"E:\WorkAndStudy\Python\config\battlepass.xlsx")
parser.add_argument('--path_config_drop', type=str, default=r"E:\WorkAndStudy\Python\config\drop.xlsx")
parser.add_argument('--path_result_battlepass', type=str, default=r"E:\WorkAndStudy\Python\result\result_battlepass.xlsx")
parser.add_argument('--a', type=int, default=0)
parser.add_argument('--b', type=int, default=0)
parser.add_argument('--input_path', type=str, default=r'E:\WorkAndStudy\Python\result-cn\result_battlepass.xlsx')
parser.add_argument('--output_path',type=str, default=r'E:\WorkAndStudy\Python\result-cn\result_battlepass-color.xlsx')

#获取参数
args = parser.parse_args()

#修改待测试BP起初ID参数
path_config_battlepass = args.path_config_battlepass
path_config_drop = args.path_config_drop
path_result_battlepass = args.path_result_battlepass
a = args.a
b = args.b
input_path = args.input_path
output_path = args.output_path

#读取battlepass
bpid = list(range(a, b+1))
data_BPLevel = pd.read_excel(path_config_battlepass, sheet_name='BPLevel')
bplevel = data_BPLevel['Id(key.uint32)*'].isin(bpid)
sheet_BPLevel = data_BPLevel[bplevel]
#免费奖励
sheet1 = sheet_BPLevel.drop(columns=['GotoType(uint32)', 'SeasonId(uint32)*', 'ImportantLevel(uint32)', 'SpecialGift(uint32)', 'FreePicRate(float)', 'PayPicRate(float)',
                                     'Exp(uint32)*', 'Id(key.uint32)*', 'PaidDrop(uint32)', 'PaidBox(uint32)', 'Cost(uint32)', 'PaidReward(array.uint32)'])
dropid_BPLevel_free = sheet1['FreeDrop(uint32)']
dropid_BPLevel_free = dropid_BPLevel_free.fillna(value='1')
dropset_BPLevel_free = [i for i in dropid_BPLevel_free.values] #if not pd.isna(i)]
dropset_BPLevel_free = [int(i) for i in dropset_BPLevel_free]
data_drop_free = pd.read_excel(path_config_drop, sheet_name='drop')
drop_columns_free = data_drop_free.columns.values
dst_free = pd.DataFrame(columns=drop_columns_free)
for dropid_free in dropset_BPLevel_free:
    oneidx_free = data_drop_free["id(key.uint32)*"] == dropid_free
    oneline_free_ = data_drop_free[oneidx_free].values
    if len(oneline_free_)>0:
        oneline_free = oneline_free_[0]
        dst_free.loc[len(dst_free)] = oneline_free
    else:
        print(dropid_free, '找不到')
sheet2 = dst_free.drop(columns=['limitRepeated(array.string)', 'showId(uint32)', 'dropPackage(array.string)', 'limitExtraDropId(uint32)', 'guaranteeDropList(array.string)',
                                'guaranteeXRound(array.string)'])

#付费奖励
sheet3 = sheet_BPLevel.drop(columns=['Id(key.uint32)*', 'SeasonId(uint32)*', 'Level(uint32)*', 'Exp(uint32)*', 'FreeDrop(uint32)', 'FreeBox(uint32)', 'Cost(uint32)', 'FreeReward(array.uint32)',
                                     'GotoType(uint32)', 'ImportantLevel(uint32)', 'SpecialGift(uint32)', 'FreePicRate(float)', 'PayPicRate(float)'])
dropid_BPLevel_paid = sheet3['PaidDrop(uint32)']
dropid_BPLevel_paid = dropid_BPLevel_paid.fillna(value='1')
dropset_BPLevel_paid = [i for i in dropid_BPLevel_paid.values if not pd.isna(i)]
dropset_BPLevel_paid = [int(i) for i in dropset_BPLevel_paid]
data_drop_paid = pd.read_excel(path_config_drop, sheet_name='drop')
drop_columns_paid = data_drop_paid.columns.values
dst_paid = pd.DataFrame(columns=drop_columns_paid)
for dropid_paid in dropset_BPLevel_paid:
    oneidx_paid = data_drop_paid["id(key.uint32)*"] == dropid_paid
    oneline_paid_ = data_drop_paid[oneidx_paid].values
    if len(oneline_paid_)>0:
        oneline_paid = oneline_paid_[0]
        dst_paid.loc[len(dst_paid)] = oneline_paid
    else:
        print(dropid_paid, '找不到')
sheet4 = dst_paid.drop(columns=['limitRepeated(array.string)', 'showId(uint32)', 'dropPackage(array.string)', 'limitExtraDropId(uint32)', 'guaranteeDropList(array.string)',
                                'guaranteeXRound(array.string)'])

#将数据写入新表
excel_writer = pd.ExcelWriter(path_result_battlepass)
sheet1.to_excel(excel_writer, sheet_name='bplevel', index=False)
sheet2.to_excel(excel_writer, sheet_name='bplevel', index=False, startcol=5)
sheet3.to_excel(excel_writer, sheet_name='bplevel', index=False, startcol=11)
sheet4.to_excel(excel_writer, sheet_name='bplevel', index=False, startcol=15)
excel_writer.save()

#对比数据
input_path = r'E:\WorkAndStudy\Python\result-cn\result_battlepass.xlsx'
output_path = r'E:\WorkAndStudy\Python\result-cn\result_battlepass-color.xlsx'
sheet = 'bplevel'
client1 = 3
server1 = 8
client2 = 13
server2 = 18

#取左边数
def ClientValLeft(strval):
    # print('str:{}, clientvalleft:{}'.format(strval, strval.split('|')[0]))
    return strval.split('|')[0]
def ServerValLeft(strval):
    # print('str:{}, servervalleft:{}'.format(strval, strval.split(':')[0]))
    return strval.split(':')[0]

#取右边数
def ClientValRight(strval):
    # print('str:{}, clientvalright:{}'.format(strval, strval.split('|')[1]))
    return strval.split('|')[1]
def ServerValRight(strval):
    # print('str:{}, servervalright:{}'.format(strval, strval.split(':')[1]))
    return strval.split(':')[1]

# def ABCto123(strval):
#     return ord(strval) - ord('A')

def color_spray(input_path, output_path, sheet, client1, server1, client2, server2):
    xlsx_data = pd.read_excel(input_path, sheet_name=sheet)
    col1_name = xlsx_data.columns[client1]
    col2_name = xlsx_data.columns[server1]
    col1_data = xlsx_data[col1_name]
    col2_data = xlsx_data[col2_name]
    col3_name = xlsx_data.columns[client2]
    col4_name = xlsx_data.columns[server2]
    col3_data = xlsx_data[col3_name]
    col4_data = xlsx_data[col4_name]

    #获取免费2列中左值不同的行
    diff_lines1 = []
    for i in range(len(col1_data.values)):
        if ClientValLeft(str(col1_data.values[i])) != ServerValLeft(str(col2_data.values[i])):
            diff_lines1.append(i)
            # print(diff_lines1)
    #获取免费2列中右值不同的行
    diff_lines2 = []
    for i in range(len(col1_data.values)):
        if ClientValRight(str(col1_data.values[i])) != ServerValRight(str(col2_data.values[i])):
            diff_lines2.append(i)
            # print(diff_lines2)
    #获取付费2列中左值不同的行
    diff_lines3 = []
    for i in range(len(col3_data.values)):
        if ClientValLeft(str(col3_data.values[i])) != ServerValLeft(str(col4_data.values[i])):
            diff_lines3.append(i)
            # print(diff_lines3)
    # 获取付费2列中右值不同的行
    diff_lines4 = []
    for i in range(len(col3_data.values)):
        if ClientValRight(str(col3_data.values[i])) != ServerValRight(str(col4_data.values[i])):
            diff_lines4.append(i)
            # print(diff_lines4)

    """上色"""
    wb = load_workbook(filename=input_path)
    # 使用第一个sheet作为工作簿
    work = wb[sheet]
    fill1 = PatternFill('solid', fgColor='fff68f')
    fill2 = PatternFill('solid', fgColor='ff715f')
    fill3 = PatternFill('solid', fgColor='fff68f')
    fill4 = PatternFill('solid', fgColor='ff715f')


    for i in diff_lines1:
        # work[i].fill = fill1
        work.cell(row=i + 2, column=client1 + 1).fill = fill1
        work.cell(row=i + 2, column=server1 + 1).fill = fill1
    for i in diff_lines2:
        # work[i].fill = fill1
        work.cell(row=i + 2, column=client1 + 1).fill = fill2
        work.cell(row=i + 2, column=server1 + 1).fill = fill2
    for i in diff_lines3:
        # work[i].fill = fill1
        work.cell(row=i + 2, column=client2 + 1).fill = fill3
        work.cell(row=i + 2, column=server2 + 1).fill = fill3
    for i in diff_lines4:
        # work[i].fill = fill1
        work.cell(row=i + 2, column=client2 + 1).fill = fill4
        work.cell(row=i + 2, column=server2 + 1).fill = fill4

    wb.close()
    wb.save(output_path)

color_spray(input_path, output_path, sheet, client1, server1, client2, server2)













