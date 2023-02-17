import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import argparse
 
#命令行解析器对象
parser = argparse.ArgumentParser(description='Demo of argparse') 
#命令行参数
parser.add_argument('--inpath', type=str, default=r"E:\WorkAndStudy\Python\result-cn\result_battlepass.xlsx")
parser.add_argument('--outpath', type=str, default=r"E:\WorkAndStudy\Python\result-cn\result_battlepass-color.xlsx")
parser.add_argument('--col1', type=str, default="D")
parser.add_argument('--col2', type=str, default="I")
parser.add_argument('--sheet', type=str, default="bplevel")

#获取参数
args = parser.parse_args()

def LeftVal(strval):
    print("str:{},leftval:{}".format(strval,strval.split(":")[0].split("|")[0]))
    return strval.split(":")[0].split("|")[0]

def ABCto123(strval):
    return ord(strval)-ord('A')

def color_spray(input_path, output_path,sheet,col1,col2):
    xlsx_data = pd.read_excel(input_path, sheet_name=sheet)
    col1_name=xlsx_data.columns[ABCto123(col1)]
    col2_name=xlsx_data.columns[ABCto123(col2)]
    
    col1_data=xlsx_data[col1_name]
    col2_data=xlsx_data[col2_name]
 
  
    #获取2个列中左值不同的行
    diff_lines=[]
    for i in range(len(col1_data.values)):
        if LeftVal(str(col1_data.values[i]))!=LeftVal(str(col2_data.values[i])):
            diff_lines.append(i)
    print(diff_lines)
    
    """上色"""
    wb = load_workbook(filename=input_path)
    # 使用第一个sheet作为工作簿
    work = wb[sheet]
    fill1 = PatternFill('solid', fgColor='fff68f')
    for i in diff_lines:
        # work[i].fill = fill1
        work.cell(row=i+2,column=ABCto123(col1)+1).fill = fill1
        work.cell(row=i+2,column=ABCto123(col2)+1).fill = fill1
    wb.close()
    wb.save(output_path)
    
    
inpath=args.inpath
outpath=args.outpath
sheet=args.sheet
col1=args.col1
col2=args.col2
color_spray(inpath, outpath,sheet,col1,col2)