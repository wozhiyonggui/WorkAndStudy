import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

#配置目录
path_config_shopstore = r'C:\Users\jinyan\PycharmProjects\Python\config\ShopStore.xlsx'
path_result_shopstore = r'C:\Users\jinyan\PycharmProjects\Python\result\result_shopstore.xlsx'
path_config_drop = r'C:\Users\jinyan\PycharmProjects\Python\config\drop.xlsx'
out_path = r'C:\Users\jinyan\PycharmProjects\Python\result\result_shopstore_color.xlsx'
sheet = '9'


#读取子表ShopStore
data = pd.read_excel(path_config_shopstore, sheet_name='ShopStore')
#读取字段Group等于9的数据
data1 = data['Group(uint32)'] == 9
sheet1 = data[data1].drop(columns=['SubGroup(uint32)', 'Sort(uint32)', 'Type(uint32)', 'Price1(array.uint32)', 'Price2(array.uint32)', 'ValidTime(int64)', 'Tag(array.string)', 'Redpoint(uint32)'
    , 'DirectTo(array.uint32)', 'DirectToTime(string)', 'HideEnd(uint32)', 'Limit(array.uint32)', 'HotGroup(uint32)', 'Count(uint32)', 'MaxLimit(uint32)', 'BoxRule(string)', 'LimitMaxNum(uint32)'
                                  , 'SaleStartTime(string)', 'SaleEndTime(string)', 'NewbieSaleTime(int64)', 'Rarity(string)', 'ExtraDropList(array.uint32)', 'PayId(uint32)', 'FirstPrice(uint32)'
    , 'ExchangePic(string)', 'AdjustSort(uint32)', 'LimitVip(uint32)', 'LimitDes(uint32)', 'DropType(uint32)', 'MergeDrop(uint32)', 'VipDirectTo(array.uint32)', 'shopID(key.uint32)'])

#从sheet1中读取需要的dropid，并且在空格中填充1
dropid = sheet1['DropId(uint32)'].fillna(value='1')
#将取出的dropid值设置成列表
dropid = [i for i in dropid.values]
#将列表中的所有内容转换成int
dropid = [int(i) for i in dropid]
#读取drop表
data_drop = pd.read_excel(path_config_drop, sheet_name='drop')
#读取drop表中所有列的名称
data_columns = data_drop.columns.values
#通过drop的列建立数据结构
dst = pd.DataFrame(columns=data_columns)
#依次取dropid从drop表中找出对应的行，并且读取所有数据
for id in dropid:
    oneidx = data_drop['id(key.uint32)*'] == id
    oneline_ = data_drop[oneidx].values
    if len(oneline_)>0:
        oneline = oneline_[0]
        dst.loc[len(dst)] = oneline
sheet2 = dst.drop(columns=['limitRepeated(array.string)', 'showId(uint32)', 'dropPackage(array.string)', 'limitExtraDropId(uint32)', 'guaranteeDropList(array.string)',
                                'guaranteeXRound(array.string)'])
#将sheet1和2写入excel表
excel_writer = pd.ExcelWriter(path_result_shopstore)
sheet1.to_excel(excel_writer, sheet_name='9', index=False)
sheet2.to_excel(excel_writer, sheet_name='9', index=False, startcol=11)
excel_writer.save()

#读取excel表
data2 = pd.read_excel(path_result_shopstore, sheet_name='9')
#读取客户端id
col1 = data2['ObjectId(uint32)']
#读取服务器id
col2 = data2['dropItem(array.string)']
#建立空列表
diff_lines = []
#根据左边客户端id数量做为次数进行循环，客户端id与服务器id的左边数值如果不相同，则将i写入列表
for i in range(len(col1.values)):
    if str(col1.values[i]) != str(col2.values[i]).split(':')[0]:
        diff_lines.append(i)

wb = load_workbook(filename=path_result_shopstore)
work = wb[sheet]
fill1 = PatternFill('solid', fgColor='fff68f')
for i in diff_lines:
    work.cell(row=i + 2, column=1+1).fill = fill1
    work.cell(row=i + 2, column=14+1).fill = fill1

wb.close()
wb.save(out_path)







